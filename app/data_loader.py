"""Load the curriculum (LOs) and book chunks from disk.

Parses LO.xlsx into structured LearningOutcome objects and chunks.json into
Chunk objects. Both are loaded once at startup and held in memory.
"""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

from app.schemas import LearningOutcome, Chunk


# Matches a leading LO code like "6.5.2.1.1" optionally preceded by
# "Learning Outcome " and followed by ":" or whitespace.
_LO_ID_RE = re.compile(r"^(?:Learning Outcome\s+)?(\d+(?:\.\d+)+)\s*:?\s*(.*)", re.DOTALL)

# Matches the domain/subdomain label like "Domain 2: ... Subdomain 2.1: ..."
_DOMAIN_RE = re.compile(
    r"(Domain\s+\d+:\s*[^.]+?)\.\s*(Subdomain\s+[\d.]+:\s*.+)",
    re.IGNORECASE,
)


def _parse_lo_text(raw: str) -> tuple[str, str]:
    """Pull the LO id and trimmed text out of a raw spreadsheet cell."""
    raw = raw.strip()
    m = _LO_ID_RE.match(raw)
    if not m:
        # Fallback: no recognizable id, use a hash-like placeholder
        return raw[:20], raw
    lo_id, text = m.group(1), m.group(2).strip()
    if not text:
        text = raw
    return lo_id, text


def _parse_domain_label(raw: str) -> tuple[str, str]:
    """Split "Domain 2: ... Subdomain 2.1: ..." into (domain, subdomain)."""
    raw = raw.strip()
    m = _DOMAIN_RE.search(raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # Fallback: best effort split on ". Subdomain"
    if ". Subdomain" in raw:
        a, b = raw.split(". Subdomain", 1)
        return a.strip(), ("Subdomain" + b).strip()
    return raw, ""


def load_learning_outcomes(xlsx_path: str | Path) -> list[LearningOutcome]:
    """Load LOs from the curriculum spreadsheet.

    Expects a sheet with header row "Domain | Learning Outcome" and
    each subsequent row giving the domain/subdomain label and the LO text.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    outcomes: list[LearningOutcome] = []

    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not row or not any(row):
            continue
        first = (row[0] or "").strip() if row[0] else ""
        # Skip preamble rows and the header
        if not header_seen:
            if first.lower().startswith("domain") and len(row) > 1 and (row[1] or "").lower().startswith("learning"):
                header_seen = True
            continue
        if len(row) < 2 or not row[0] or not row[1]:
            continue
        domain_label = str(row[0])
        lo_raw = str(row[1])
        lo_id, lo_text = _parse_lo_text(lo_raw)
        domain, subdomain = _parse_domain_label(domain_label)
        outcomes.append(
            LearningOutcome(
                lo_id=lo_id,
                text=lo_text,
                domain=domain,
                subdomain=subdomain,
                full_domain_label=domain_label,
            )
        )
    return outcomes


def load_chunks(json_path: str | Path) -> list[Chunk]:
    """Load pre-chunked book content."""
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    chunks: list[Chunk] = []
    for item in raw:
        page_span = item.get("pageSpan") or {}
        chunks.append(
            Chunk(
                chunk_id=item["chunkId"],
                content=item.get("content", ""),
                page_start=page_span.get("pageStart"),
                page_end=page_span.get("pageEnd"),
            )
        )
    return chunks
